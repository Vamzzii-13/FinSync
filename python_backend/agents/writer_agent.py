from openpyxl import Workbook
from openpyxl.styles import Alignment

def writer_agent(data: list, file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Report"

    # Reordered headers as requested
    headers = [
        "S.No.", "Vendor/Shop Name", "Date", "GSTIN", "Invoice No.", 
        "HSN Codes", "CGST", "SGST", "IGST", "Total Tax", "Taxable Amount"
    ]

    # Set optimal column widths for proper cell fitting
    column_widths = {
        'A': 8,   # S.No.
        'B': 35,  # Vendor/Shop Name (wider for multi-line)
        'C': 15,  # Date
        'D': 18,  # GSTIN
        'E': 20,  # Invoice No.
        'F': 45,  # HSN Codes (wider for multiple codes)
        'G': 12,  # CGST
        'H': 12,  # SGST
        'I': 12,  # IGST
        'J': 12,  # Total Tax
        'K': 15   # Taxable Amount
    }

    # Apply column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Create simple headers with center alignment
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.alignment = Alignment(vertical='center', horizontal='center')

    # Process data rows
    for row_idx, record in enumerate(data, start=2):
        # Serial number
        cell = ws.cell(row=row_idx, column=1)
        cell.value = row_idx - 1
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Handle multi-line shop name properly
        shop_name = record.get("Shop Name", "N/A")
        hsn_line_count = 1  # Default for HSN codes
        if isinstance(shop_name, str):
            # Keep line breaks but clean them up
            shop_name = shop_name.replace('\\n', '\n').strip()
            # Count lines to determine row height
            line_count = shop_name.count('\n') + 1
        else:
            line_count = 1
            
        # Pre-calculate HSN codes for row height calculation
        items = record.get("Items", [])
        if items and isinstance(items, list):
            hsn_codes = []
            for item in items:
                if isinstance(item, dict) and "HSN Code" in item:
                    hsn_code = str(item["HSN Code"]).strip()
                    if hsn_code and hsn_code != "0" and hsn_code != "null":
                        hsn_codes.append(hsn_code)
            
            if hsn_codes:
                unique_codes = list(dict.fromkeys(hsn_codes))
                if len(unique_codes) <= 2:
                    hsn_line_count = 1
                elif len(unique_codes) <= 6:
                    hsn_line_count = (len(unique_codes) + 1) // 2  # Groups of 2
                else:
                    hsn_line_count = (len(unique_codes) + 2) // 3  # Groups of 3
        
        # Vendor/Shop Name with text wrapping (Column B)
        cell = ws.cell(row=row_idx, column=2)
        cell.value = shop_name
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Date (Column C)
        cell = ws.cell(row=row_idx, column=3)
        date_value = record.get("Invoice Date", "N/A")
        if date_value in (None, "", "null", 0, 0.0):
            date_value = "N/A"
        cell.value = date_value
        cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # GSTIN (Column D)
        cell = ws.cell(row=row_idx, column=4)
        gstin_value = record.get("GSTIN", "N/A")
        if gstin_value in (None, "", "null", 0, 0.0):
            gstin_value = "N/A"
        cell.value = gstin_value
        cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # Invoice No. (Column E)
        cell = ws.cell(row=row_idx, column=5)
        invoice_value = record.get("Invoice Number", "N/A")
        if invoice_value in (None, "", "null", 0, 0.0):
            invoice_value = "N/A"
        cell.value = invoice_value
        cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # HSN Codes (Column F) - get from Items array
        items = record.get("Items", [])
        cell = ws.cell(row=row_idx, column=6)
        
        if items and isinstance(items, list):
            # Extract HSN codes from Items array
            hsn_codes = []
            for item in items:
                if isinstance(item, dict) and "HSN Code" in item:
                    hsn_code = str(item["HSN Code"]).strip()
                    if hsn_code and hsn_code != "0" and hsn_code != "null":
                        hsn_codes.append(hsn_code)
            
            if hsn_codes:
                # Remove duplicates while preserving order
                unique_codes = list(dict.fromkeys(hsn_codes))
                
                # Format for better readability - max 2 codes per line
                if len(unique_codes) <= 2:
                    cell.value = ", ".join(unique_codes)
                elif len(unique_codes) <= 6:
                    # Split into groups of 2 for medium lists
                    formatted_codes = []
                    for i in range(0, len(unique_codes), 2):
                        chunk = unique_codes[i:i+2]
                        formatted_codes.append(", ".join(chunk))
                    cell.value = "\n".join(formatted_codes)
                else:
                    # For large lists, group by 3 per line
                    formatted_codes = []
                    for i in range(0, len(unique_codes), 3):
                        chunk = unique_codes[i:i+3]
                        formatted_codes.append(", ".join(chunk))
                    cell.value = "\n".join(formatted_codes)
            else:
                cell.value = "N/A"
        else:
            # Fallback to direct HSN Code field
            hsn_codes = record.get("HSN Code", "N/A")
            cell.value = hsn_codes if hsn_codes else "N/A"
            
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Tax columns: CGST, SGST, IGST, Total Tax (Columns G, H, I, J)
        tax_data = [
            record.get("CGST", "N/A"),
            record.get("SGST", "N/A"),
            record.get("IGST", "N/A"),
            record.get("Tax Amount", "N/A")
        ]
        
        for col_idx, value in enumerate(tax_data, start=7):
            cell = ws.cell(row=row_idx, column=col_idx)
            if value in (None, "", "null", 0, 0.0):
                value = "N/A"
            cell.value = value
            cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # Taxable Amount (Column K)
        cell = ws.cell(row=row_idx, column=11)
        taxable_value = record.get("Total Amount", "N/A")
        if taxable_value in (None, "", "null", 0, 0.0):
            taxable_value = "N/A"
        cell.value = taxable_value
        cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # Set row height based on content (consider both shop name and HSN codes)
        max_lines = max(line_count, hsn_line_count)
        dynamic_height = max(25, 15 + (max_lines * 15))
        ws.row_dimensions[row_idx].height = dynamic_height

    # Set header row height
    ws.row_dimensions[1].height = 25

    wb.save(file_path)
    print(f"📝 Excel report saved to: {file_path}")